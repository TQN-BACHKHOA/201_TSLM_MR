import queue, math, os, subprocess
from openpyxl import load_workbook
from sympy import *
from os import path
import latexFile as ltx
init_printing()

wb = load_workbook("data.xlsx")
ws = wb.worksheets[0]
average_length = 0 
entropy = 0
efficiency = 0
freq = []
probability_CS = 0
huffman_string = []
output_freq = []

if not path.exists("output"):
    os.mkdir("output")
path1 = 'output'
file = 'latex_output.tex'
with open(os.path.join(path1, file), 'w') as fp: 
    pass
latexFile = ltx.latexClass(os.path.join(path1, file))

class HuffmanNode(object):
    def __init__(self, left=None, right=None):
        self.left = left
        self.right = right
    def children(self):
        return (self.left, self.right)
    def preorder(self, path=None):
        global average_length
        if path is None:                                
            path = []
        if self.left is not None:
            if isinstance(myMapFunc(self.left[1]), HuffmanNode):   
                latexFile.addNode_tree(_text=convert(path+['0']), _type="NODE")
                myMapFunc(self.left[1]).preorder(path+['0'])
                latexFile.closeNode_tree()
            else:
                latexFile.addNode_tree(_text=convert(self.left[1]+convert([': ']+path+['0'])), _type="LEAF")
                output_freq.append((self.left[1], str(self.left[0]), convert(path+['0'])))
                print(convert(self.left[1] + convert([': ']+path+['0'])))
                average_length += self.left[0]*(len(path)+1)
                latexFile.closeNode_tree()
            
        if self.right is not None:
            if isinstance(myMapFunc(self.right[1]), HuffmanNode):
                latexFile.addNode_tree(_text=convert(path+['1']), _type="NODE")
                myMapFunc(self.right[1]).preorder(path+['1'])
                latexFile.closeNode_tree()
            else:
                latexFile.addNode_tree(_text=convert(self.right[1]+convert([': ']+path+['1'])), _type="LEAF")
                output_freq.append((self.right[1], str(self.right[0]), convert(path+['1'])))
                print(convert(self.right[1] + convert([': ']+path+['1'])))
                average_length += self.right[0]*(len(path)+1)
                latexFile.closeNode_tree()

def encode(frequencies):
    p = queue.PriorityQueue()
    for item in frequencies:
        p.put(item)
    while p.qsize() > 1:
        Left, Right = p.get(), p.get()
        node = HuffmanNode(Left, Right)
        huffman_string.append((node, str(node)))
        p.put((Left[0] + Right[0], str(node)))
    return p.get()

def convert(char_array):
    return("".join(char_array))

def myMapFunc(_string=None):
    for _node in huffman_string:
        if _string == _node[1]:
            return _node[0]
    return _string

## Begin the program    

for row in ws.iter_rows(min_col=1, max_col=2, min_row=2):
    row = [cell.value for cell in row]
    freq.append((row[0], str(row[1])))
    probability_CS += row[0]
if round(probability_CS, 5) == 1:
    print("Data satisfied")
else:
    print("Data not satisfied")
    os._exit(0)

for item in freq:
    entropy += item[0]*math.log2(1/item[0])
node = encode(freq)

###Write to tex file
latexFile.init()
rootNODE = myMapFunc(node[1])
latexFile.init_tree()
rootNODE.preorder()
latexFile.end_tree()
latexFile.add_figure("The Huffman Tree")

latexFile.add_table(_array=output_freq)
latexFile.add_figure("Bảng từ mã các nodes")

i, H, N = symbols('i H N')

latexFile.addString(input_string="Entropy:")
entropy_exp = Eq(Sum(Indexed('p',i)*log(1/(Indexed('p',i)),2),(i,0,'k')), round(entropy,3))
latexFile.addString(input_string=latex(entropy_exp), math=True)

latexFile.addString(input_string="Chiều dài TB từ mã:")
tuma_exp = Eq(Sum(Indexed('p',i)*Indexed('N',i),(i,0,'k')), round(average_length,3))
latexFile.addString(input_string='N = '+latex(tuma_exp), math=True)

latexFile.addString(input_string="Hiệu suất Huffman:")
efficiency = entropy/average_length
efficiency_exp = Eq((H/N), round(efficiency,3))
latexFile.addString(input_string=latex(efficiency_exp), math=True)

latexFile.end()

###Convert from tex to pdf
x = subprocess.call('pdflatex -output-directory=output output/latex_output.tex', shell=True)